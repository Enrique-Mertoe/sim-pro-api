"""
Team management RPC functions
"""
from datetime import timezone, timedelta
from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from django.utils import timezone
from ..models import User, Team, SimCard, LotMetadata

SSMAuthUser = get_user_model()


def get_my_team_details(user):
    """Get detailed information about the user's team"""
    try:
        ssm_user = user
        if not ssm_user.team:
            return {'error': 'User is not assigned to a team'}

        team = ssm_user.team

        # Get team members with their sim card counts
        team_members = User.objects.filter(team=team).select_related('auth_user')
        members_data = []

        for member in team_members:
            sim_count = SimCard.objects.filter(assigned_to_user=member).count()
            active_sim_count = SimCard.objects.filter(assigned_to_user=member, status='active').count()

            members_data.append({
                'id': str(member.id),
                'name': f"{member.auth_user.first_name} {member.auth_user.last_name}".strip(),
                'email': member.auth_user.email,
                'role': member.role,
                'sim_card_count': sim_count,
                'active_sim_count': active_sim_count,
                'is_active': member.is_active,
                'joined_at': member.created_at.isoformat() if member.created_at else None,
            })

        # Get team statistics
        total_sim_cards = SimCard.objects.filter(assigned_to_user__team=team).count()
        active_sim_cards = SimCard.objects.filter(assigned_to_user__team=team, status='active').count()

        return {
            'team': {
                'id': str(team.id),
                'name': team.name,
                'created_at': team.created_at.isoformat() if team.created_at else None,
                'member_count': len(members_data),
                'total_sim_cards': total_sim_cards,
                'active_sim_cards': active_sim_cards,
            },
            'members': members_data,
            'user_role': ssm_user.role,
            'is_team_leader': ssm_user.role == 'team_leader'
        }
    except User.DoesNotExist:
        return {'error': 'User profile not found'}


def get_team_performance(user, team_id=None):
    """Get team performance metrics"""
    try:
        ssm_user = User.objects.get(auth_user=user)

        # If team_id is provided, check if user has permission to view it
        if team_id:
            if ssm_user.role not in ['admin', 'team_leader']:
                raise PermissionError("Insufficient permissions to view team performance")
            team = Team.objects.get(id=team_id)
        else:
            # Use user's own team
            if not ssm_user.team:
                return {'error': 'User is not assigned to a team'}
            team = ssm_user.team

        # Calculate performance metrics
        team_members = User.objects.filter(team=team)
        performance_data = []

        for member in team_members:
            total_assigned = SimCard.objects.filter(assigned_to=member).count()
            active_cards = SimCard.objects.filter(assigned_to=member, status='active').count()
            inactive_cards = SimCard.objects.filter(assigned_to=member, status='inactive').count()

            performance_data.append({
                'member_id': str(member.id),
                'member_name': f"{member.auth_user.first_name} {member.auth_user.last_name}".strip(),
                'total_assigned': total_assigned,
                'active_cards': active_cards,
                'inactive_cards': inactive_cards,
                'activation_rate': round((active_cards / total_assigned * 100) if total_assigned > 0 else 0, 2)
            })

        return {
            'team_id': str(team.id),
            'team_name': team.name,
            'performance': performance_data,
            'generated_at': user.date_joined.isoformat() if hasattr(user, 'date_joined') else None
        }
    except User.DoesNotExist:
        return {'error': 'User profile not found'}
    except Team.DoesNotExist:
        return {'error': 'Team not found'}


def get_available_teams(user):
    """Get list of all available teams (admin only)"""
    try:
        ssm_user = User.objects.get(auth_user=user)
        if ssm_user.role != 'admin':
            raise PermissionError("Only admins can view all teams")

        teams = Team.objects.annotate(member_count=Count('user')).all()

        return [
            {
                'id': str(team.id),
                'name': team.name,
                'member_count': team.member_count,
                'created_at': team.created_at.isoformat() if team.created_at else None,
            }
            for team in teams
        ]
    except User.DoesNotExist:
        raise PermissionError("User profile not found")


def get_teams_analytics(user):
    """Get analytics data for all teams (admin/team_leader only)"""
    try:
        ssm_user = User.objects.get(auth_user=user)
        if ssm_user.role not in ['admin', 'team_leader']:
            raise PermissionError("Only admins and team leaders can view team analytics")

        # Filter teams based on role
        if ssm_user.role == 'team_leader':
            # Team leaders can only see their own team
            teams_query = Team.objects.filter(id=ssm_user.team.id) if ssm_user.team else Team.objects.none()
        else:
            # Admins can see all teams
            teams_query = Team.objects.all()

        teams_analytics = []

        for team in teams_query.select_related('leader'):
            # Get team members
            team_members = User.objects.filter(team=team)

            # Get SIM card counts for the team
            team_sim_cards = SimCard.objects.filter(assigned_to_user__team=team)

            total_sim_cards = team_sim_cards.count()
            quality_sim_cards = team_sim_cards.filter(quality='quality').count()
            non_quality_sim_cards = team_sim_cards.filter(quality='non_quality').count()
            registered_sim_cards = team_sim_cards.filter(registered_on__isnull=False).count()
            active_sim_cards = team_sim_cards.filter(status='active').count()

            # Calculate rates
            quality_rate = round((quality_sim_cards / total_sim_cards * 100) if total_sim_cards > 0 else 0, 2)
            registration_rate = round((registered_sim_cards / total_sim_cards * 100) if total_sim_cards > 0 else 0, 2)

            teams_analytics.append({
                'team_id': str(team.id),
                'team_name': team.name,
                'region': team.region,
                'territory': team.territory,
                'leader': {
                    'id': str(team.leader.id) if team.leader else None,
                    'name': f"{team.leader.auth_user.first_name} {team.leader.auth_user.last_name}".strip() if team.leader else None,
                    'email': team.leader.auth_user.email if team.leader else None,
                } if team.leader else None,
                'member_count': team_members.count(),
                'is_active': team.is_active,
                'created_at': team.created_at.isoformat() if team.created_at else None,
                'sim_card_stats': {
                    'total_sim_cards': total_sim_cards,
                    'quality_sim_cards': quality_sim_cards,
                    'non_quality_sim_cards': non_quality_sim_cards,
                    'registered_sim_cards': registered_sim_cards,
                    'active_sim_cards': active_sim_cards,
                    'quality_rate': quality_rate,
                    'registration_rate': registration_rate,
                }
            })

        # Sort by total SIM cards (descending)
        teams_analytics.sort(key=lambda x: x['sim_card_stats']['total_sim_cards'], reverse=True)

        # Calculate overall summary
        total_teams = len(teams_analytics)
        total_sim_cards_all = sum(t['sim_card_stats']['total_sim_cards'] for t in teams_analytics)
        total_quality_all = sum(t['sim_card_stats']['quality_sim_cards'] for t in teams_analytics)
        total_registered_all = sum(t['sim_card_stats']['registered_sim_cards'] for t in teams_analytics)

        return {
            'teams': teams_analytics,
            'summary': {
                'total_teams': total_teams,
                'active_teams': len([t for t in teams_analytics if t['is_active']]),
                'total_sim_cards': total_sim_cards_all,
                'total_quality_sim_cards': total_quality_all,
                'total_registered_sim_cards': total_registered_all,
                'overall_quality_rate': round(
                    (total_quality_all / total_sim_cards_all * 100) if total_sim_cards_all > 0 else 0, 2),
                'overall_registration_rate': round(
                    (total_registered_all / total_sim_cards_all * 100) if total_sim_cards_all > 0 else 0, 2),
            }
        }

    except User.DoesNotExist:
        raise PermissionError("User profile not found")


def get_groups_summery(user):
    if user.role not in ['admin', 'team_leader']:
        raise PermissionError("Access denied")

    from ssm.models.base_models import TeamGroup, TeamGroupMembership
    from django.db.models import Count, Q

    # Get user's team
    team = user.team
    if not team:
        return {"groups": [], "total_groups": 0}

    # Get groups with member counts and activity stats
    groups = TeamGroup.objects.filter(
        team=team,
        is_active=True
    ).annotate(
        member_count=Count('memberships', distinct=True),
        active_member_count=Count(
            'memberships__user',
            filter=Q(memberships__user__is_active=True),
            distinct=True
        )
    ).order_by('-member_count', 'name')

    groups_data = []
    for group in groups:
        # Get recent activity (you can customize this based on your needs)
        recent_activity = group.memberships.filter(
            joined_at__gte=timezone.now() - timedelta(days=30)
        ).count()

        groups_data.append({
            "id": str(group.id),
            "name": group.name,
            "description": group.description or "",
            "location": group.location or "",
            "member_count": group.member_count,
            "active_member_count": group.active_member_count,
            "recent_activity": recent_activity,
            "created_at": group.created_at.isoformat(),
        })

    return {
        "groups": groups_data,
        "total_groups": len(groups_data)
    }


def team_overview_data(user):
    """Get team overview data with member performance metrics"""
    if user.role not in ['admin', 'team_leader']:
        raise PermissionError("Access denied")

    # Get user's team
    team = user.team
    if not team:
        return {"error": "User is not assigned to a team", "team_members": []}

    # Get team members with their performance data
    team_members = (User.objects.filter(team=team)
                    .exclude(
        role='team_leader'
    ).select_related('auth_user'))
    members_data = []

    for member in team_members:
        # Get SIM card statistics
        member_sim_cards = SimCard.objects.filter(assigned_to_user=member)
        total_sim_cards = member_sim_cards.count()
        quality_cards = member_sim_cards.filter(quality='quality').count()
        active_cards = member_sim_cards.filter(status='active').count()

        # Calculate quality rate
        quality_rate = round((quality_cards / total_sim_cards * 100) if total_sim_cards > 0 else 0, 1)

        # Calculate performance score (based on quality rate and activity)
        # This is a simplified calculation - you can make it more sophisticated
        activity_bonus = 1.0 if member.is_active else 0.5
        performance_score = round((quality_rate / 10) * activity_bonus, 1)

        # Get last activity (using last updated SIM card or created_at as fallback)
        last_sim_activity = member_sim_cards.order_by('-updated_at').first()
        last_activity = last_sim_activity.updated_at if last_sim_activity else member.created_at

        # Determine status
        status = "active" if member.is_active else "inactive"

        # Get member's location (you might need to adjust this based on your model structure)
        location = getattr(member, 'location', '') or getattr(team, 'region', 'N/A')

        members_data.append({
            "id": str(member.id),
            "name": f"{member.auth_user.first_name} {member.auth_user.last_name}".strip() or member.auth_user.email,
            "email": member.auth_user.email,
            "phone": member.phone_number or "N/A",
            "role": member.role.replace('_', ' ').title() if member.role else "Field Agent",
            "location": location,
            "status": status,
            "sim_cards_assigned": total_sim_cards,
            "quality_cards": quality_cards,
            "quality_rate": quality_rate,
            "last_activity": last_activity.strftime("%Y-%m-%d %H:%M") if last_activity else "N/A",
            "performance_score": performance_score
        })

    # Sort by performance score descending
    members_data.sort(key=lambda x: x['performance_score'], reverse=True)

    return {
        "team_members": members_data,
        "team_stats": {
            "total_members": len(members_data),
            "active_members": len([m for m in members_data if m['status'] == 'active']),
            "total_sim_cards": sum(m['sim_cards_assigned'] for m in members_data),
            "average_quality_rate": round(
                sum(m['quality_rate'] for m in members_data) / len(members_data) if members_data else 0, 1),
            "average_performance": round(
                sum(m['performance_score'] for m in members_data) / len(members_data) if members_data else 0, 1)
        }
    }


def team_allocation(user):
    """Get all teams with their allocation breakdown"""
    try:
        ssm_user = user
        if ssm_user.role not in ['admin']:
            raise PermissionError("Only admins and team leaders can view team allocation")

        teams = Team.objects.filter(admin=user, is_default=False).all()
        allocation_data = []

        for team in teams:
            # team_lots = LotMetadata.objects.filter(
            #     assigned_team=team,
            #     admin=user
            # )
            team_sim_cards = SimCard.objects.filter(
                # serial_number__in=team_lots.serial_numbers,
                team=team,
                admin=user
            )

            print("allo", team_sim_cards)
            total_allocated = team_sim_cards.count()
            assigned = team_sim_cards.filter(
                assigned_to_user__isnull=False
            ).count()
            unassigned = team_sim_cards.filter(
                assigned_to_user__isnull=True
            ).count()
            stock = team_sim_cards.filter(
                activation_date__isnull=True
            ).count()

            allocation_data.append({
                'team_id': str(team.id),
                'team_name': team.name,
                'region': team.region or 'N/A',
                'territory': team.territory or 'N/A',
                'leader': {
                    'id': str(team.leader.id) if team.leader else None,
                    'name': f"{team.leader.full_name}".strip() if team.leader else 'No Leader',
                    'email': team.leader.auth_user.email if team.leader else None,
                } if team.leader else None,
                'allocated': total_allocated,
                'assigned': assigned,
                'unassigned': unassigned,
                'stock': stock,
                'is_active': team.is_active
            })

        return {
            'teams': allocation_data,
            'total_teams': len(allocation_data),
            'total_allocated': sum(t['allocated'] for t in allocation_data),
            'total_assigned': sum(t['assigned'] for t in allocation_data),
            'total_unassigned': sum(t['unassigned'] for t in allocation_data)
        }

    except User.DoesNotExist:
        raise PermissionError("User profile not found")


def normalize_quality(value):
    if value in ['Y', 'QUALITY']:
        return 'Y'
    elif value in ['N', 'NON_QUALITY']:
        return 'N'
    return 'N/A'


def export_team_allocation_excel(user):
    """Export team allocation to Excel with multiple sheets and colors"""
    try:
        if user.role not in ['admin']:
            raise PermissionError("Only admins can export team allocation")
        import pandas as pd
        import base64
        from io import BytesIO
        from openpyxl.styles import PatternFill
        teams = Team.objects.filter(admin=user)

        if not teams.exists():
            raise Exception("No teams found for export")

        output = BytesIO()
        all_data = []
        team_colors = ['FFE6E6', 'E6F3FF', 'E6FFE6', 'FFF0E6', 'F0E6FF', 'FFFFE6', 'E6FFFF', 'FFE6F0']

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Collect all data
            for team in teams:
                sim_cards = SimCard.objects.filter(team=team, admin=user).values(
                    'serial_number', 'quality', 'assigned_to_user__full_name',
                    'ba_msisdn', 'mobigo', 'activation_date', 'sale_date', 'top_up_amount'
                )

                for sim in sim_cards:
                    all_data.append({
                        'Team': team.name,
                        'Region': team.region or 'N/A',
                        'Serial Number': sim['serial_number'],
                        'Quality': normalize_quality(sim['quality']),
                        'Assigned To': sim['assigned_to_user__full_name'] or 'Unassigned',
                        'BA MSISDN': sim['ba_msisdn'] or 'N/A',
                        'Mobigo': sim['mobigo'] or 'N/A',
                        'Activation Date': sim['activation_date'].replace(tzinfo=None) if sim[
                            'activation_date'] else None,
                        'Sale Date': sim['sale_date'].replace(tzinfo=None) if sim['sale_date'] else None,
                        'Top Up Amount': sim['top_up_amount'] or 0
                    })

            # Create general sheet (always create at least one sheet)
            if all_data:
                df_all = pd.DataFrame(all_data)
            else:
                df_all = pd.DataFrame([{'Team': 'No Data', 'Serial Number': 'No SIM cards found'}])

            df_all.to_excel(writer, sheet_name='All Teams', index=False)

            # Apply colors and filters only if we have real data
            if all_data:
                worksheet = writer.sheets['All Teams']
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Set column widths
                worksheet.column_dimensions['C'].width = 25  # Serial Number
                worksheet.column_dimensions['F'].width = 18  # BA MSISDN
                worksheet.column_dimensions['G'].width = 18  # Mobigo
                worksheet.column_dimensions['H'].width = 20  # Activation Date
                worksheet.column_dimensions['I'].width = 20  # Sale Date
                
                current_team = None
                color_index = -1

                for row_idx, row in enumerate(df_all.itertuples(), start=2):
                    if row.Team != current_team:
                        current_team = row.Team
                        color_index = (color_index + 1) % len(team_colors)

                    fill = PatternFill(start_color=team_colors[color_index],
                                       end_color=team_colors[color_index],
                                       fill_type='solid')

                    for col in range(1, len(df_all.columns) + 1):
                        worksheet.cell(row=row_idx, column=col).fill = fill

            # Individual team sheets
            for team in teams:
                sim_cards = SimCard.objects.filter(team=team, admin=user).values(
                    'serial_number', 'quality', 'assigned_to_user__full_name',
                    'ba_msisdn', 'mobigo', 'activation_date', 'sale_date', 'top_up_amount', 'created_at'
                )

                team_data = []
                for sim in sim_cards:
                    team_data.append({
                        'Serial Number': sim['serial_number'],
                        'Quality': normalize_quality(sim['quality']),
                        'Assigned To': sim['assigned_to_user__full_name'] or 'Unassigned',
                        'BA MSISDN': sim['ba_msisdn'] or 'N/A',
                        'Mobigo': sim['mobigo'] or 'N/A',
                        'Activation Date': sim['activation_date'].replace(tzinfo=None) if sim[
                            'activation_date'] else None,
                        'Sale Date': sim['sale_date'].replace(tzinfo=None) if sim['sale_date'] else None,
                        'Top Up Amount': sim['top_up_amount'] or 0,
                        'Created Date': sim['created_at'].replace(tzinfo=None) if sim['created_at'] else None
                    })

                if team_data:
                    df_team = pd.DataFrame(team_data)
                    sheet_name = team.name[:31]
                    df_team.to_excel(writer, sheet_name=sheet_name, index=False)
                    team_worksheet = writer.sheets[sheet_name]
                    team_worksheet.auto_filter.ref = team_worksheet.dimensions
                    
                    # Set column widths for team sheets
                    team_worksheet.column_dimensions['A'].width = 25  # Serial Number
                    team_worksheet.column_dimensions['D'].width = 18  # BA MSISDN
                    team_worksheet.column_dimensions['E'].width = 18  # Mobigo
                    team_worksheet.column_dimensions['F'].width = 20  # Activation Date
                    team_worksheet.column_dimensions['G'].width = 20  # Sale Date
                    team_worksheet.column_dimensions['I'].width = 20  # Created Date

        output.seek(0)
        excel_data = output.getvalue()
        base64_data = base64.b64encode(excel_data).decode('utf-8')

        return {
            'success': True,
            'file_base64': base64_data,
            'filename': f'team_allocation_{user.full_name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'total_teams': teams.count(),
            'total_sim_cards': len(all_data)
        }

    except Exception as e:
        raise Exception(f'Failed to export team allocation: {str(e)}')


# Register functions
functions = {
    'get_my_team_details': get_my_team_details,
    'get_team_performance': get_team_performance,
    'get_available_teams': get_available_teams,
    'get_teams_analytics': get_teams_analytics,
    'get_groups_summery': get_groups_summery,
    'team_overview_data': team_overview_data,
    'team_allocation': team_allocation,
    'export_team_allocation_excel': export_team_allocation_excel,
}
